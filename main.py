"""
공공데이터포털: (신)동네예보정보조회서비스 API
공공데이터포털: (신)생활기상지수조회 API
xlsx 파일 분석 후 해당 값을 python 데이터로 변환.

The latest modify day: Thu, December 18, 2018
- from Leni
"""
#파일을 읽어들이는 내장함수인 load_workbook을 불러옵니다.
from openpyxl import load_workbook
#파일이름이 angel.xlsx인 파일을 불러옵니다.
angelEx=load_workbook(filename='file/location.xlsx')
#불러온 엑셀 파일 중 데이터를 찾을 sheet의 이름을 입력합니다.
sheet1 = angelEx['Sheet1']

city_location = {}

for city in sheet1.iter_rows(min_row=2):
    level1 = city[0].value
    if city_location.get(level1) is None and level1 is not None:
        xyCode = {
            'x':int(city[3].value),
            'y':int(city[4].value),
            'code':int(city[5].value)
            }
        city_location[level1] = [xyCode, {}]

    elif city_location.get(level1) is not None:

        level2 = city[1].value
        if city_location[level1][1].get(level2) is None and level2 is not None:
            xyCode = {
                'x':int(city[3].value),
                'y':int(city[4].value),
                'code':int(city[5].value)
                }
            city_location[level1][1][level2] = [xyCode, {}]

        elif city_location[level1][1].get(level2) is not None:

            level3 = city[2].value
            if city_location[level1][1][level2][1].get(level3) is None and level3 is not None:
                xyCode = {
                    'x':int(city[3].value),
                    'y':int(city[4].value),
                    'code':int(city[5].value)
                    }
                city_location[level1][1][level2][1][level3] = xyCode

print(city_location)
