"""
공공데이터포털: (신)동네예보정보조회서비스 API
공공데이터포털: (신)생활기상지수조회 API
xlsx 파일 분석 후 해당 값을 python 데이터로 변환.

The latest modify day: Thu, December 18, 2018
- from Leni
"""
#파일을 읽어들이는 내장함수인 load_workbook을 불러옵니다.
from openpyxl import load_workbook
#파일이름이 angel.xlsx인 파일을 불러옵니다.

def make_data():
    city_location = {}
    angelEx=load_workbook(filename='file/location.xlsx')
    #불러온 엑셀 파일 중 데이터를 찾을 sheet의 이름을 입력합니다.
    sheet1 = angelEx['Sheet1']
    for city in sheet1.iter_rows(min_row=2):
        make_city(city_location, city, level=0)
    return city_location

def make_city(temp_location, row, level):
    level_city = row[level].value
    if temp_location.get(level_city) is None and level_city is not None:
        temp_location[level_city] = make_xyCode(level, row[3].value, row[4].value, row[5].value)

    elif temp_location.get(level_city) is not None:
        level += 1
        make_city(temp_location[level_city][1], row, level)


def make_xyCode(level, x, y, code):
    xyCode = {
        'x':int(x),
        'y':int(y),
        'code':int(code)
        }

    if level == 2:
        return xyCode
    else:
        return [xyCode, {}]
